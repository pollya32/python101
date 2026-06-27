import io
import base64
from flask import Flask, request, jsonify, send_file, render_template
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB


def preprocess_image(img):
    """이미지 전처리: 대비 향상 및 노이즈 제거"""
    img = img.convert('L')  # 흑백 변환
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    enhancer = ImageEnhance.Sharpness(img)
    img = enhancer.enhance(2.0)
    return img


def cluster_positions(values, tolerance):
    """가까운 위치값들을 같은 행/열로 묶기"""
    if not values:
        return []
    sorted_vals = sorted(set(values))
    clusters = [[sorted_vals[0]]]
    for v in sorted_vals[1:]:
        if v - clusters[-1][-1] <= tolerance:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return [int(sum(c) / len(c)) for c in clusters]


def extract_table_data(img):
    """이미지에서 테이블 데이터 추출"""
    processed = preprocess_image(img)

    # 한국어 + 영어 OCR (없으면 영어만)
    try:
        data = pytesseract.image_to_data(
            processed,
            lang='kor+eng',
            output_type=pytesseract.Output.DICT,
            config='--psm 6'
        )
    except pytesseract.TesseractError:
        data = pytesseract.image_to_data(
            processed,
            lang='eng',
            output_type=pytesseract.Output.DICT,
            config='--psm 6'
        )

    # 신뢰도 50 이상인 텍스트만 사용
    words = []
    for i in range(len(data['text'])):
        text = data['text'][i].strip()
        conf = int(data['conf'][i])
        if text and conf >= 50:
            words.append({
                'text': text,
                'left': data['left'][i],
                'top': data['top'][i],
                'width': data['width'][i],
                'height': data['height'][i],
                'center_x': data['left'][i] + data['width'][i] // 2,
                'center_y': data['top'][i] + data['height'][i] // 2,
            })

    if not words:
        return []

    # 행/열 클러스터링
    avg_height = sum(w['height'] for w in words) / len(words)
    row_tolerance = max(avg_height * 0.6, 8)

    y_positions = [w['center_y'] for w in words]
    row_centers = cluster_positions(y_positions, row_tolerance)

    x_positions = [w['center_x'] for w in words]
    avg_width = sum(w['width'] for w in words) / len(words)
    col_tolerance = max(avg_width * 1.5, 20)
    col_centers = cluster_positions(x_positions, col_tolerance)

    # 그리드에 단어 배치
    grid = {}
    for w in words:
        row_idx = min(range(len(row_centers)), key=lambda i: abs(row_centers[i] - w['center_y']))
        col_idx = min(range(len(col_centers)), key=lambda i: abs(col_centers[i] - w['center_x']))
        key = (row_idx, col_idx)
        if key in grid:
            grid[key] += ' ' + w['text']
        else:
            grid[key] = w['text']

    if not grid:
        return []

    max_row = max(k[0] for k in grid.keys()) + 1
    max_col = max(k[1] for k in grid.keys()) + 1

    table = []
    for r in range(max_row):
        row = []
        for c in range(max_col):
            row.append(grid.get((r, c), ''))
        table.append(row)

    return table


def create_excel(table_data):
    """테이블 데이터를 Excel 파일로 변환"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR 결과"

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    col_widths = {}

    for r_idx, row in enumerate(table_data, 1):
        for c_idx, cell_value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = alt_fill

            col_letter = get_column_letter(c_idx)
            text_len = len(str(cell_value)) if cell_value else 0
            col_widths[col_letter] = max(col_widths.get(col_letter, 8), text_len + 4)

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(width, 40)

    ws.row_dimensions[1].height = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/ocr', methods=['POST'])
def ocr():
    try:
        if 'image' in request.files:
            file = request.files['image']
            img = Image.open(file.stream)
        elif request.is_json and 'image_data' in request.json:
            image_data = request.json['image_data']
            if ',' in image_data:
                image_data = image_data.split(',')[1]
            img_bytes = base64.b64decode(image_data)
            img = Image.open(io.BytesIO(img_bytes))
        else:
            return jsonify({'error': '이미지가 없습니다'}), 400

        table_data = extract_table_data(img)

        if not table_data:
            return jsonify({'error': '텍스트를 인식하지 못했습니다. 이미지를 확인해 주세요.'}), 422

        # TSV 생성 (Excel 붙여넣기용)
        tsv_lines = []
        for row in table_data:
            tsv_lines.append('\t'.join(str(cell) for cell in row))
        tsv = '\n'.join(tsv_lines)

        return jsonify({
            'table': table_data,
            'tsv': tsv,
            'rows': len(table_data),
            'cols': len(table_data[0]) if table_data else 0
        })

    except Exception as e:
        return jsonify({'error': f'처리 중 오류 발생: {str(e)}'}), 500


@app.route('/download', methods=['POST'])
def download():
    try:
        data = request.json
        table_data = data.get('table', [])

        if not table_data:
            return jsonify({'error': '데이터가 없습니다'}), 400

        excel_file = create_excel(table_data)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='OCR_결과.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
