"""
반도체 공정 DOE (Design of Experiments) 분석 시스템
목표: 두께(Thickness), 균일도(Uniformity), 파티클(Particle) 최적화
변수: AR, N2, HE, C3H6, O2, NF3, PRESSURE, SPACE, TIME, RF_POWER
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import rcParams
import seaborn as sns
from itertools import combinations
from scipy import stats
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression, RidgeCV
from sklearn.metrics import r2_score
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
import warnings
import os
import json
from datetime import datetime

warnings.filterwarnings('ignore')

# ── 한글 폰트 설정 ──────────────────────────────────────────────────────────
def setup_korean_font():
    font_candidates = [
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    ]
    for path in font_candidates:
        if os.path.exists(path):
            fe = fm.FontEntry(fname=path, name='KoreanFont')
            fm.fontManager.ttflist.insert(0, fe)
            rcParams['font.family'] = 'KoreanFont'
            return
    rcParams['font.family'] = 'DejaVu Sans'

setup_korean_font()
rcParams['axes.unicode_minus'] = False


# ── 상수 정의 ───────────────────────────────────────────────────────────────
FACTORS = {
    'AR':       {'unit': 'sccm', 'low': 0,    'high': 100,  'type': 'gas'},
    'N2':       {'unit': 'sccm', 'low': 0,    'high': 200,  'type': 'gas'},
    'HE':       {'unit': 'sccm', 'low': 0,    'high': 500,  'type': 'gas'},
    'C3H6':     {'unit': 'sccm', 'low': 0,    'high': 50,   'type': 'gas'},
    'O2':       {'unit': 'sccm', 'low': 0,    'high': 100,  'type': 'gas'},
    'NF3':      {'unit': 'sccm', 'low': 0,    'high': 100,  'type': 'gas'},
    'PRESSURE': {'unit': 'Torr', 'low': 1,    'high': 10,   'type': 'process'},
    'SPACE':    {'unit': 'mm',   'low': 5,    'high': 30,   'type': 'process'},
    'TIME':     {'unit': 'sec',  'low': 10,   'high': 120,  'type': 'process'},
    'RF_POWER': {'unit': 'W',    'low': 100,  'high': 2000, 'type': 'process'},
}

RESPONSES = {
    'THICKNESS':    {'unit': 'Å',    'target': 5000, 'target_type': 'nominal', 'weight': 0.4},
    'UNIFORMITY':   {'unit': '%',    'target': 1.0,  'target_type': 'smaller', 'weight': 0.4},
    'PARTICLE':     {'unit': 'ea',   'target': 0,    'target_type': 'smaller', 'weight': 0.2},
}

OUTPUT_DIR = 'doe_results'

# ── 300mm 웨이퍼 25포인트 계측 좌표 (mm, 중심 기준) ──────────────────────────
WAFER_RADIUS = 150.0  # mm (300mm 웨이퍼)
WAFER_EDGE_EXCLUSION = 3.0  # mm

# 배치: 중심 1 + r=35(4) + r=75(4) + r=105(4) + r=120(4) + r=140(8) = 25점
WAFER_POINTS_25 = [
    # P01: 중심
    (  0.0,    0.0),
    # P02-P05: r=35mm, 상하좌우
    ( 35.0,    0.0), (  0.0,  35.0), (-35.0,   0.0), (  0.0, -35.0),
    # P06-P09: r=75mm, 대각
    ( 53.0,   53.0), (-53.0,  53.0), (-53.0, -53.0), ( 53.0, -53.0),
    # P10-P13: r=105mm, 상하좌우
    (105.0,    0.0), (  0.0, 105.0), (-105.0,  0.0), (  0.0,-105.0),
    # P14-P17: r=120mm, 대각
    ( 84.9,   84.9), (-84.9,  84.9), (-84.9, -84.9), ( 84.9, -84.9),
    # P18-P25: r=140mm, 8방향
    (140.0,    0.0), ( 99.0,  99.0), (  0.0, 140.0), (-99.0,  99.0),
    (-140.0,   0.0), (-99.0, -99.0), (  0.0,-140.0), ( 99.0, -99.0),
]
POINT_LABELS = [f'P{i+1:02d}' for i in range(25)]


# ══════════════════════════════════════════════════════════════════════════════
#  DOE 실험 설계 생성기
# ══════════════════════════════════════════════════════════════════════════════
class DOEDesigner:
    """실험 계획 설계 클래스"""

    def __init__(self, factors: dict, active_factors: list = None):
        self.factors = factors
        self.active_factors = active_factors or list(factors.keys())
        self.k = len(self.active_factors)

    def _encode(self, name: str, value: float) -> float:
        lo, hi = self.factors[name]['low'], self.factors[name]['high']
        return 2 * (value - lo) / (hi - lo) - 1

    def _decode(self, name: str, coded: float) -> float:
        lo, hi = self.factors[name]['low'], self.factors[name]['high']
        return lo + (coded + 1) / 2 * (hi - lo)

    def full_factorial_2level(self) -> pd.DataFrame:
        """2수준 완전 요인 설계"""
        n = 2 ** self.k
        design = np.zeros((n, self.k))
        for i, name in enumerate(self.active_factors):
            period = 2 ** (self.k - i - 1)
            pattern = np.tile(
                np.repeat([-1, 1], period), n // (2 * period)
            )
            design[:, i] = pattern
        df = pd.DataFrame(design, columns=self.active_factors)
        for name in self.active_factors:
            df[f'{name}_actual'] = df[name].apply(
                lambda x: self._decode(name, x)
            )
        df['Run'] = range(1, n + 1)
        return df[['Run'] + self.active_factors +
                  [f'{n}_actual' for n in self.active_factors]]

    def box_behnken(self) -> pd.DataFrame:
        """Box-Behnken 설계 (k=3인 경우)"""
        if self.k < 3:
            raise ValueError("Box-Behnken requires at least 3 factors")
        k = self.k
        pairs = list(combinations(range(k), 2))
        rows = []
        for i, j in pairs:
            for si, sj in [(-1,-1),(-1,1),(1,-1),(1,1)]:
                row = [0.0] * k
                row[i] = si
                row[j] = sj
                rows.append(row)
        # 중심점 3회
        for _ in range(3):
            rows.append([0.0] * k)
        design = np.array(rows)
        df = pd.DataFrame(design, columns=self.active_factors)
        for name in self.active_factors:
            df[f'{name}_actual'] = df[name].apply(
                lambda x: self._decode(name, x)
            )
        df['Run'] = range(1, len(df) + 1)
        return df[['Run'] + self.active_factors +
                  [f'{n}_actual' for n in self.active_factors]]

    def central_composite(self, alpha: float = 1.414) -> pd.DataFrame:
        """중심 합성 설계 (CCD)"""
        k = self.k
        # 2^k 완전 요인부
        factorial_runs = 2 ** k
        design_factorial = np.array(
            [[((-1) ** ((i >> j) & 1)) for j in range(k)]
             for i in range(factorial_runs)]
        )
        # 성형점(star points)
        star = np.zeros((2 * k, k))
        for i in range(k):
            star[2*i, i] = alpha
            star[2*i+1, i] = -alpha
        # 중심점
        center = np.zeros((4, k))
        design = np.vstack([design_factorial, star, center])
        df = pd.DataFrame(design, columns=self.active_factors)
        for name in self.active_factors:
            df[f'{name}_actual'] = df[name].apply(
                lambda x: self._decode(name, x)
            )
        df['Run'] = range(1, len(df) + 1)
        return df[['Run'] + self.active_factors +
                  [f'{n}_actual' for n in self.active_factors]]

    def plackett_burman(self, n_runs: int = 12) -> pd.DataFrame:
        """Plackett-Burman 스크리닝 설계 (최대 11인자)"""
        pb_generators = {
            12: [1,-1,1,1,1,-1,-1,-1,1,-1,-1],
            20: [1,1,-1,-1,1,1,1,1,-1,1,-1,1,-1,-1,-1,-1,1,1,-1],
        }
        if n_runs not in pb_generators:
            n_runs = 12
        gen = pb_generators[n_runs]
        matrix = []
        row = gen[:]
        for _ in range(n_runs - 1):
            matrix.append(row[:])
            row = [row[-1]] + row[:-1]
        matrix.append([-1] * (n_runs - 1))
        design = np.array(matrix)[:, :self.k]
        df = pd.DataFrame(design, columns=self.active_factors)
        for name in self.active_factors:
            df[f'{name}_actual'] = df[name].apply(
                lambda x: self._decode(name, x)
            )
        df['Run'] = range(1, len(df) + 1)
        return df[['Run'] + self.active_factors +
                  [f'{n}_actual' for n in self.active_factors]]


# ══════════════════════════════════════════════════════════════════════════════
#  통계 분석기
# ══════════════════════════════════════════════════════════════════════════════
class DOEAnalyzer:
    """DOE 결과 통계 분석 클래스"""

    def __init__(self, data: pd.DataFrame, factors: list, responses: list):
        self.data = data.copy()
        self.factors = factors
        self.responses = responses

    def _build_formula(self, response: str, interactions: int = 2) -> str:
        """statsmodels 회귀식 공식 생성 (자유도 자동 조정)"""
        n_obs = len(self.data)
        terms = list(self.factors)
        # 교호작용 포함 시 자유도 확인
        if interactions >= 2:
            pairs = list(combinations(self.factors, 2))
            if len(terms) + len(pairs) + 1 < n_obs:
                for a, b in pairs:
                    terms.append(f'{a}:{b}')
        if interactions >= 3:
            triples = list(combinations(self.factors, 3))
            if len(terms) + len(triples) + 1 < n_obs:
                for a, b, c in triples:
                    terms.append(f'{a}:{b}:{c}')
        return f'{response} ~ {" + ".join(terms)}'

    def anova(self, response: str, interactions: int = 2) -> pd.DataFrame:
        """ANOVA 분석 (자유도 부족 시 주효과만)"""
        for lvl in [interactions, 1, 0]:
            formula = self._build_formula(response, lvl)
            try:
                model = ols(formula, data=self.data).fit()
                table = anova_lm(model, typ=2)
                table['Significant'] = table['PR(>F)'] < 0.05
                return table
            except Exception:
                continue
        return pd.DataFrame()

    def regression(self, response: str, degree: int = 2) -> dict:
        """다항 회귀 분석 (관측수 부족 시 Ridge 정규화 자동 적용)"""
        X = self.data[self.factors].values
        y = self.data[response].values
        n_obs = len(y)
        # 관측수 대비 차수 자동 조정
        actual_degree = degree
        for d in range(degree, 0, -1):
            test_poly = PolynomialFeatures(degree=d, include_bias=True)
            n_features = test_poly.fit_transform(X[:1]).shape[1]
            if n_features <= n_obs:
                actual_degree = d
                break
        poly = PolynomialFeatures(degree=actual_degree, include_bias=False)
        Xp = poly.fit_transform(X)
        n_features = Xp.shape[1]
        # 과적합 구간: Ridge 정규화 사용
        if n_features >= n_obs - 2:
            alphas = np.logspace(-3, 4, 50)
            model = RidgeCV(alphas=alphas, cv=min(5, n_obs - 1)).fit(Xp, y)
        else:
            model = LinearRegression().fit(Xp, y)
        y_pred = model.predict(Xp)
        ss_res = np.sum((y - y_pred) ** 2)
        ss_tot = np.sum((y - np.mean(y)) ** 2)
        r2 = 1 - ss_res / (ss_tot + 1e-12)
        # intercept를 별도 보관 (predict에 쓰임)
        return {
            'model': model,
            'poly': poly,
            'r2': float(r2),
            'degree': actual_degree,
            'y_pred': y_pred,
            'feature_names': poly.get_feature_names_out(self.factors),
            'coefficients': getattr(model, 'coef_', np.zeros(n_features)),
        }

    def main_effects(self, response: str) -> pd.DataFrame:
        """주효과 계산"""
        results = []
        for f in self.factors:
            if f not in self.data.columns:
                continue
            groups = self.data.groupby(f)[response].mean()
            if len(groups) >= 2:
                effect = groups.max() - groups.min()
                t_stat, p_val = stats.ttest_ind(
                    self.data[self.data[f] == self.data[f].min()][response],
                    self.data[self.data[f] == self.data[f].max()][response],
                )
                results.append({
                    'Factor': f,
                    'Effect': effect,
                    'Low_Mean': groups.min(),
                    'High_Mean': groups.max(),
                    'T-stat': t_stat,
                    'P-value': p_val,
                    'Significant': p_val < 0.05,
                })
        return pd.DataFrame(results).sort_values('Effect', key=abs, ascending=False)

    def interaction_effects(self, response: str) -> pd.DataFrame:
        """2-인자 교호작용 분석"""
        results = []
        for a, b in combinations(self.factors, 2):
            if a not in self.data.columns or b not in self.data.columns:
                continue
            try:
                a_vals = sorted(self.data[a].unique())
                b_vals = sorted(self.data[b].unique())
                if len(a_vals) < 2 or len(b_vals) < 2:
                    continue
                a_lo, a_hi = a_vals[0], a_vals[-1]
                b_lo, b_hi = b_vals[0], b_vals[-1]
                def mean_at(av, bv):
                    mask = (self.data[a] == av) & (self.data[b] == bv)
                    s = self.data[mask][response]
                    return s.mean() if len(s) > 0 else np.nan
                ypp = mean_at(a_hi, b_hi)
                ypm = mean_at(a_hi, b_lo)
                ymp = mean_at(a_lo, b_hi)
                ymm = mean_at(a_lo, b_lo)
                if any(np.isnan(v) for v in [ypp, ypm, ymp, ymm]):
                    continue
                interaction = ((ypp - ypm) - (ymp - ymm)) / 2
                results.append({
                    'Factors': f'{a}*{b}',
                    'A': a, 'B': b,
                    'Interaction_Effect': interaction,
                    'AB_HiHi': ypp, 'AB_HiLo': ypm,
                    'AB_LoHi': ymp, 'AB_LoLo': ymm,
                })
            except Exception:
                continue
        df = pd.DataFrame(results)
        if not df.empty:
            df = df.sort_values('Interaction_Effect', key=abs, ascending=False)
        return df

    def pareto_effects(self, response: str) -> pd.DataFrame:
        """파레토 효과 차트용 데이터"""
        me = self.main_effects(response)
        ie = self.interaction_effects(response)
        rows = []
        for _, r in me.iterrows():
            rows.append({'Term': r['Factor'], 'Effect': abs(r['Effect']),
                         'Signed': r['Effect'], 'Type': 'Main'})
        for _, r in ie.iterrows():
            rows.append({'Term': r['Factors'], 'Effect': abs(r['Interaction_Effect']),
                         'Signed': r['Interaction_Effect'], 'Type': 'Interaction'})
        df = pd.DataFrame(rows).sort_values('Effect', ascending=False)
        total = df['Effect'].sum()
        df['Cumulative%'] = (df['Effect'].cumsum() / total * 100) if total > 0 else 0
        return df


# ══════════════════════════════════════════════════════════════════════════════
#  최적화 엔진
# ══════════════════════════════════════════════════════════════════════════════
class DOEOptimizer:
    """다중 반응 최적화 (Desirability Function)"""

    def __init__(self, analyzer: DOEAnalyzer, responses_config: dict):
        self.analyzer = analyzer
        self.responses_config = responses_config
        self.models = {}
        self._fit_models()

    def _fit_models(self):
        for resp in self.analyzer.responses:
            if resp in self.analyzer.data.columns:
                self.models[resp] = self.analyzer.regression(resp, degree=2)

    def _desirability(self, response: str, value: float) -> float:
        cfg = self.responses_config[response]
        target = cfg['target']
        rtype = cfg['target_type']
        clamp = lambda v: float(np.clip(v, 0.0, 1.0))
        if rtype == 'nominal':
            tol = max(target * 0.15, 1.0)
            return clamp(1 - abs(value - target) / tol)
        elif rtype == 'smaller':
            if response in self.analyzer.data.columns:
                worst = self.analyzer.data[response].max()
            else:
                worst = target * 10
            worst = max(worst, target + 1e-9)
            if value <= target:
                return 1.0
            return clamp(1 - (value - target) / (worst - target))
        elif rtype == 'larger':
            if response in self.analyzer.data.columns:
                worst = self.analyzer.data[response].min()
            else:
                worst = 0
            worst = min(worst, target - 1e-9)
            if value >= target:
                return 1.0
            return clamp((value - worst) / (target - worst))
        return 0.5

    def overall_desirability(self, factor_values: dict) -> float:
        X = np.array([[factor_values.get(f, 0) for f in self.analyzer.factors]])
        d_vals = []
        weights = []
        for resp, mdl in self.models.items():
            Xp = mdl['poly'].transform(X)
            pred = float(mdl['model'].predict(Xp)[0])
            # 데이터 범위 내로 클리핑
            if resp in self.analyzer.data.columns:
                lo = self.analyzer.data[resp].min()
                hi = self.analyzer.data[resp].max()
                pred = float(np.clip(pred, max(0, lo * 0.5), hi * 1.5))
            d = float(np.clip(self._desirability(resp, pred), 0.0, 1.0))
            w = self.responses_config[resp]['weight']
            d_vals.append(d)
            weights.append(w)
        if not d_vals or sum(weights) == 0:
            return 0.0
        # Weighted geometric mean of desirabilities
        total_w = sum(weights)
        log_d = sum(w * np.log(max(d, 1e-12)) for d, w in zip(d_vals, weights))
        return float(np.clip(np.exp(log_d / total_w), 0.0, 1.0))

    def optimize(self, n_random: int = 2000) -> dict:
        """랜덤 + 국소 최적화로 최적 조건 탐색"""
        from scipy.optimize import minimize

        bounds = []
        factor_list = self.analyzer.factors
        for f in factor_list:
            if f in FACTORS:
                bounds.append((FACTORS[f]['low'], FACTORS[f]['high']))
            else:
                col = self.analyzer.data[f]
                bounds.append((col.min(), col.max()))

        def neg_desirability(x):
            fv = dict(zip(factor_list, x))
            return -self.overall_desirability(fv)

        # 1) 랜덤 탐색으로 유망 시작점 찾기
        rng = np.random.default_rng(42)
        best_x, best_d = None, -1
        for _ in range(n_random):
            x0 = np.array([rng.uniform(lo, hi) for lo, hi in bounds])
            d = -neg_desirability(x0)
            if d > best_d:
                best_d = d
                best_x = x0.copy()

        # 2) 국소 최적화 (L-BFGS-B)
        result = minimize(neg_desirability, best_x,
                          method='L-BFGS-B', bounds=bounds,
                          options={'maxiter': 500, 'ftol': 1e-9})
        opt_x = result.x if result.success else best_x
        best_params = dict(zip(factor_list, opt_x))
        best_desirability = -neg_desirability(opt_x)

        predictions = {}
        X = np.array([[best_params[f] for f in factor_list]])
        for resp, mdl in self.models.items():
            Xp = mdl['poly'].transform(X)
            pred = float(mdl['model'].predict(Xp)[0])
            # 물리적으로 불가능한 값 클리핑 (데이터 범위 기준)
            if resp in self.analyzer.data.columns:
                lo = self.analyzer.data[resp].min()
                hi = self.analyzer.data[resp].max()
                pred = float(np.clip(pred, max(0, lo * 0.5), hi * 1.5))
            predictions[resp] = pred

        return {
            'optimal_factors': best_params,
            'desirability': float(best_desirability),
            'predicted_responses': predictions,
        }

    def grid_search(self, n_grid: int = 5) -> dict:
        """grid_search는 optimize()로 위임 (하위 호환)"""
        return self.optimize(n_random=max(1000, n_grid ** min(len(self.analyzer.factors), 4)))


# ══════════════════════════════════════════════════════════════════════════════
#  시각화
# ══════════════════════════════════════════════════════════════════════════════
class DOEVisualizer:

    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.saved_files = []

    def _save(self, fig, name: str):
        path = os.path.join(self.output_dir, name)
        fig.savefig(path, dpi=120, bbox_inches='tight')
        plt.close(fig)
        self.saved_files.append(path)
        print(f"  저장: {path}")

    # ── 1. 실험 설계 행렬 ────────────────────────────────────────────────────
    def plot_design_matrix(self, design_df: pd.DataFrame, factors: list, title: str = 'Design Matrix'):
        coded_cols = [c for c in factors if c in design_df.columns]
        if not coded_cols:
            return
        data = design_df[coded_cols]
        fig, ax = plt.subplots(figsize=(max(8, len(coded_cols) * 1.2),
                                        max(6, len(data) * 0.4 + 1)))
        cmap = plt.cm.RdYlGn
        im = ax.imshow(data.values, cmap=cmap, aspect='auto', vmin=-1.5, vmax=1.5)
        ax.set_xticks(range(len(coded_cols)))
        ax.set_xticklabels(coded_cols, fontsize=9, rotation=45, ha='right')
        ax.set_yticks(range(len(data)))
        ax.set_yticklabels([f'Run {i+1}' for i in range(len(data))], fontsize=8)
        for r in range(len(data)):
            for c, col in enumerate(coded_cols):
                v = data.iloc[r, c]
                ax.text(c, r, f'{v:+.0f}', ha='center', va='center', fontsize=7,
                        color='black' if abs(v) < 1.2 else 'white')
        plt.colorbar(im, ax=ax, label='Coded Level')
        ax.set_title(title, fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, 'design_matrix.png')

    # ── 2. 주효과 도표 ───────────────────────────────────────────────────────
    def plot_main_effects(self, data: pd.DataFrame, factors: list, response: str):
        valid = [f for f in factors if f in data.columns]
        n = len(valid)
        if n == 0:
            return
        ncols = min(3, n)
        nrows = (n + ncols - 1) // ncols
        fig, axes = plt.subplots(nrows, ncols,
                                 figsize=(ncols * 4, nrows * 3.5))
        axes = np.array(axes).flatten() if n > 1 else [axes]
        for i, f in enumerate(valid):
            ax = axes[i]
            means = data.groupby(f)[response].mean()
            ax.plot(means.index, means.values, 'bo-', linewidth=2, markersize=7)
            ax.axhline(data[response].mean(), color='red', linestyle='--',
                       alpha=0.6, label='Overall Mean')
            ax.set_xlabel(f'{f} (coded)', fontsize=9)
            ax.set_ylabel(response, fontsize=9)
            ax.set_title(f'{f} Main Effect', fontsize=10, fontweight='bold')
            ax.grid(True, alpha=0.3)
            ax.legend(fontsize=7)
        for j in range(i + 1, len(axes)):
            axes[j].set_visible(False)
        fig.suptitle(f'Main Effects Plot - {response}', fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, f'main_effects_{response}.png')

    # ── 3. 교호작용 도표 ─────────────────────────────────────────────────────
    def plot_interactions(self, data: pd.DataFrame, factors: list, response: str,
                          top_n: int = 6):
        pairs = list(combinations(factors, 2))
        valid_pairs = [(a, b) for a, b in pairs
                       if a in data.columns and b in data.columns][:top_n]
        if not valid_pairs:
            return
        n = len(valid_pairs)
        ncols = min(3, n)
        nrows = (n + ncols - 1) // ncols
        fig, axes = plt.subplots(nrows, ncols,
                                 figsize=(ncols * 4.5, nrows * 3.8))
        axes = np.array(axes).flatten() if n > 1 else [axes]
        colors = ['#2196F3', '#F44336', '#4CAF50', '#FF9800']
        for idx, (a, b) in enumerate(valid_pairs):
            ax = axes[idx]
            b_vals = sorted(data[b].unique())
            for ci, bv in enumerate(b_vals):
                mask = data[b] == bv
                means = data[mask].groupby(a)[response].mean()
                ax.plot(means.index, means.values,
                        color=colors[ci % len(colors)],
                        marker='o', linewidth=2, markersize=6,
                        label=f'{b}={bv:+.0f}')
            ax.set_xlabel(f'{a} (coded)', fontsize=9)
            ax.set_ylabel(response, fontsize=9)
            ax.set_title(f'{a} × {b}', fontsize=10, fontweight='bold')
            ax.legend(fontsize=7)
            ax.grid(True, alpha=0.3)
        for j in range(idx + 1, len(axes)):
            axes[j].set_visible(False)
        fig.suptitle(f'Interaction Effects Plot - {response}',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, f'interactions_{response}.png')

    # ── 4. 파레토 차트 ───────────────────────────────────────────────────────
    def plot_pareto(self, pareto_df: pd.DataFrame, response: str, alpha: float = 0.05):
        if pareto_df.empty:
            return
        top = pareto_df.head(15)
        fig, ax1 = plt.subplots(figsize=(12, 5))
        colors = ['#E53935' if t == 'Main' else '#1E88E5' for t in top['Type']]
        bars = ax1.bar(range(len(top)), top['Effect'], color=colors, edgecolor='black',
                       linewidth=0.5, alpha=0.85)
        ax1.set_xticks(range(len(top)))
        ax1.set_xticklabels(top['Term'], rotation=45, ha='right', fontsize=9)
        ax1.set_ylabel('|Effect|', fontsize=10)
        ax1.set_xlabel('Term', fontsize=10)
        ax2 = ax1.twinx()
        ax2.plot(range(len(top)), top['Cumulative%'], 'k-o',
                 linewidth=2, markersize=5)
        ax2.axhline(80, color='gray', linestyle='--', alpha=0.5, label='80%')
        ax2.set_ylabel('Cumulative %', fontsize=10)
        ax2.set_ylim(0, 105)
        from matplotlib.patches import Patch
        legend_elements = [Patch(facecolor='#E53935', label='Main Effect'),
                           Patch(facecolor='#1E88E5', label='Interaction')]
        ax1.legend(handles=legend_elements, loc='center right', fontsize=9)
        ax1.set_title(f'Pareto Chart of Effects - {response}',
                      fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, f'pareto_{response}.png')

    # ── 5. 등고선 / 반응표면 ─────────────────────────────────────────────────
    def plot_response_surface(self, analyzer: DOEAnalyzer, response: str,
                              factor_x: str, factor_y: str,
                              fixed_values: dict = None):
        if response not in analyzer.data.columns:
            return
        mdl = analyzer.regression(response, degree=2)
        fixed = fixed_values or {f: 0 for f in analyzer.factors
                                  if f not in (factor_x, factor_y)}
        xi = np.linspace(-1, 1, 40)
        yi = np.linspace(-1, 1, 40)
        XX, YY = np.meshgrid(xi, yi)
        Z = np.zeros_like(XX)
        for r in range(XX.shape[0]):
            for c in range(XX.shape[1]):
                row = [fixed.get(f, 0) for f in analyzer.factors]
                fx_idx = analyzer.factors.index(factor_x)
                fy_idx = analyzer.factors.index(factor_y)
                row[fx_idx] = XX[r, c]
                row[fy_idx] = YY[r, c]
                Xp = mdl['poly'].transform([row])
                Z[r, c] = mdl['model'].predict(Xp)[0]

        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        # Contour
        cp = axes[0].contourf(XX, YY, Z, levels=20, cmap='RdYlGn')
        axes[0].contour(XX, YY, Z, levels=10, colors='black', linewidths=0.4, alpha=0.5)
        plt.colorbar(cp, ax=axes[0], label=f'{response}')
        axes[0].set_xlabel(f'{factor_x} (coded)', fontsize=10)
        axes[0].set_ylabel(f'{factor_y} (coded)', fontsize=10)
        axes[0].set_title(f'Contour Plot\n{response} vs {factor_x} & {factor_y}',
                          fontsize=11, fontweight='bold')
        # 3D Surface
        ax3d = fig.add_subplot(1, 2, 2, projection='3d')
        surf = ax3d.plot_surface(XX, YY, Z, cmap='RdYlGn', alpha=0.8, edgecolor='none')
        ax3d.set_xlabel(factor_x, fontsize=9)
        ax3d.set_ylabel(factor_y, fontsize=9)
        ax3d.set_zlabel(response, fontsize=9)
        ax3d.set_title(f'3D Surface Plot\n{response}', fontsize=10, fontweight='bold')
        fig.colorbar(surf, ax=ax3d, shrink=0.5, label=response)
        axes[1].remove()
        fig.tight_layout()
        self._save(fig, f'surface_{response}_{factor_x}_{factor_y}.png')

    # ── 6. 잔차 분석 ─────────────────────────────────────────────────────────
    def plot_residuals(self, analyzer: DOEAnalyzer, response: str):
        if response not in analyzer.data.columns:
            return
        mdl = analyzer.regression(response, degree=2)
        y = analyzer.data[response].values
        y_pred = mdl['y_pred']
        residuals = y - y_pred
        fig, axes = plt.subplots(2, 2, figsize=(11, 8))
        # Normal probability
        (osm, osr), (slope, intercept, r) = stats.probplot(residuals, dist='norm')
        axes[0, 0].plot(osm, osr, 'bo', markersize=5, alpha=0.7)
        axes[0, 0].plot(osm, slope * np.array(osm) + intercept, 'r-', linewidth=2)
        axes[0, 0].set_xlabel('Theoretical Quantiles')
        axes[0, 0].set_ylabel('Sample Quantiles')
        axes[0, 0].set_title('Normal Probability Plot')
        axes[0, 0].grid(True, alpha=0.3)
        # Residuals vs Fitted
        axes[0, 1].scatter(y_pred, residuals, color='steelblue', alpha=0.7, edgecolors='black', linewidths=0.3)
        axes[0, 1].axhline(0, color='red', linestyle='--')
        axes[0, 1].set_xlabel('Fitted Values')
        axes[0, 1].set_ylabel('Residuals')
        axes[0, 1].set_title('Residuals vs Fitted')
        axes[0, 1].grid(True, alpha=0.3)
        # Histogram
        axes[1, 0].hist(residuals, bins=max(5, len(residuals)//3),
                        color='steelblue', edgecolor='black', alpha=0.8)
        axes[1, 0].set_xlabel('Residuals')
        axes[1, 0].set_ylabel('Frequency')
        axes[1, 0].set_title('Residual Distribution')
        axes[1, 0].grid(True, alpha=0.3)
        # Run order
        axes[1, 1].plot(range(1, len(residuals)+1), residuals,
                        'bo-', markersize=5, alpha=0.7)
        axes[1, 1].axhline(0, color='red', linestyle='--')
        axes[1, 1].set_xlabel('Run Order')
        axes[1, 1].set_ylabel('Residuals')
        axes[1, 1].set_title('Residuals vs Run Order')
        axes[1, 1].grid(True, alpha=0.3)
        fig.suptitle(f'Residual Analysis - {response} (R²={mdl["r2"]:.3f})',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, f'residuals_{response}.png')

    # ── 7. 최적화 결과 ───────────────────────────────────────────────────────
    def plot_optimization_result(self, opt_result: dict, responses_config: dict):
        preds = opt_result['predicted_responses']
        params = opt_result['optimal_factors']
        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        # Response achievement bars
        resp_names = list(preds.keys())
        targets = [responses_config[r]['target'] for r in resp_names]
        pred_vals = [preds[r] for r in resp_names]
        x = np.arange(len(resp_names))
        w = 0.35
        bars1 = axes[0].bar(x - w/2, targets, w, label='Target', color='#4CAF50', alpha=0.8)
        bars2 = axes[0].bar(x + w/2, pred_vals, w, label='Predicted', color='#2196F3', alpha=0.8)
        axes[0].set_xticks(x)
        axes[0].set_xticklabels(resp_names, fontsize=10)
        axes[0].set_ylabel('Value', fontsize=10)
        axes[0].set_title(f'Optimal Prediction\n(Desirability={opt_result["desirability"]:.3f})',
                          fontsize=11, fontweight='bold')
        axes[0].legend(fontsize=9)
        axes[0].grid(True, alpha=0.3, axis='y')
        for bar in bars1:
            h = bar.get_height()
            axes[0].text(bar.get_x() + bar.get_width()/2, h * 1.01,
                         f'{h:.1f}', ha='center', va='bottom', fontsize=8)
        for bar in bars2:
            h = bar.get_height()
            axes[0].text(bar.get_x() + bar.get_width()/2, h * 1.01,
                         f'{h:.1f}', ha='center', va='bottom', fontsize=8, color='#1565C0')
        # Optimal factor values (radar-like bar)
        factor_names = list(params.keys())
        norm_vals = []
        for f, v in params.items():
            if f in FACTORS:
                lo, hi = FACTORS[f]['low'], FACTORS[f]['high']
                norm_vals.append((v - lo) / (hi - lo + 1e-9) * 100)
            else:
                norm_vals.append(50.0)
        colors_f = ['#E53935' if FACTORS.get(f, {}).get('type') == 'gas' else '#1E88E5'
                    for f in factor_names]
        axes[1].barh(factor_names, norm_vals, color=colors_f, edgecolor='black',
                     linewidth=0.5, alpha=0.85)
        axes[1].set_xlabel('Normalized Value (%)', fontsize=10)
        axes[1].set_title('Optimal Factor Settings\n(0%=Low, 100%=High)',
                          fontsize=11, fontweight='bold')
        axes[1].axvline(50, color='gray', linestyle='--', alpha=0.5)
        axes[1].set_xlim(0, 105)
        for i, (f, v) in enumerate(params.items()):
            unit = FACTORS.get(f, {}).get('unit', '')
            axes[1].text(norm_vals[i] + 1, i, f'{v:.1f} {unit}',
                         va='center', fontsize=8)
        from matplotlib.patches import Patch
        legend_el = [Patch(facecolor='#E53935', label='Gas'),
                     Patch(facecolor='#1E88E5', label='Process')]
        axes[1].legend(handles=legend_el, loc='lower right', fontsize=9)
        axes[1].grid(True, alpha=0.3, axis='x')
        fig.tight_layout()
        self._save(fig, 'optimization_result.png')

    # ── 8. 웨이퍼 맵 ─────────────────────────────────────────────────────────
    def plot_wafer_maps(self, result_df: pd.DataFrame, title_prefix: str = ''):
        """
        전체 DOE Run 의 웨이퍼 두께 맵을 그리드로 시각화.
        scipy interpolation 으로 웨이퍼 면을 채우고 25포인트 위치를 표시.
        """
        from scipy.interpolate import griddata

        pt_cols = [f'T_{lbl}' for lbl in POINT_LABELS]
        if not all(c in result_df.columns for c in pt_cols):
            return

        n_runs = len(result_df)
        ncols = min(4, n_runs)
        nrows = (n_runs + ncols - 1) // ncols
        fig, axes = plt.subplots(nrows, ncols,
                                 figsize=(ncols * 3.8, nrows * 3.5))
        axes = np.array(axes).flatten() if n_runs > 1 else [axes]

        pts = np.array(WAFER_POINTS_25)
        xi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 200)
        yi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 200)
        XI, YI = np.meshgrid(xi, yi)
        mask = XI**2 + YI**2 > WAFER_RADIUS**2   # 웨이퍼 밖 마스크

        # 전체 런의 두께 범위 통일
        all_vals = result_df[pt_cols].values.flatten()
        vmin, vmax = np.nanpercentile(all_vals, 2), np.nanpercentile(all_vals, 98)

        for idx, (_, row) in enumerate(result_df.iterrows()):
            ax = axes[idx]
            vals = np.array([row[c] for c in pt_cols], dtype=float)
            run_no = int(row.get('Run', idx + 1))

            # 보간 (cubic → nearest fallback)
            try:
                ZI = griddata(pts, vals, (XI, YI), method='cubic')
            except Exception:
                ZI = griddata(pts, vals, (XI, YI), method='nearest')
            ZI[mask] = np.nan

            im = ax.imshow(ZI, extent=[-WAFER_RADIUS, WAFER_RADIUS,
                                        -WAFER_RADIUS, WAFER_RADIUS],
                           origin='lower', cmap='RdYlGn',
                           vmin=vmin, vmax=vmax, aspect='equal')
            # 웨이퍼 외곽선
            circle = plt.Circle((0, 0), WAFER_RADIUS,
                                 fill=False, color='black', linewidth=1.5)
            ax.add_patch(circle)
            # 25포인트 표시
            ax.scatter(pts[:, 0], pts[:, 1], c='black', s=18, zorder=5)
            # 균일도 계산
            unif_val = (vals.max() - vals.min()) / (2 * vals.mean() + 1e-9) * 100
            mean_t = vals.mean()
            ax.set_title(f'Run {run_no}\n'
                         f'Mean={mean_t:.0f}Å  UNI={unif_val:.1f}%',
                         fontsize=8.5, fontweight='bold')
            ax.set_xlim(-WAFER_RADIUS * 1.05, WAFER_RADIUS * 1.05)
            ax.set_ylim(-WAFER_RADIUS * 1.05, WAFER_RADIUS * 1.05)
            ax.set_xticks([])
            ax.set_yticks([])
            plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

        for j in range(idx + 1, len(axes)):
            axes[j].set_visible(False)

        fig.suptitle(f'{title_prefix}300mm 웨이퍼 두께 맵 (25포인트)',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, 'wafer_maps_all_runs.png')

    def plot_wafer_best_worst(self, result_df: pd.DataFrame):
        """최고/최저 균일도 Run 의 웨이퍼 맵 비교"""
        from scipy.interpolate import griddata

        pt_cols = [f'T_{lbl}' for lbl in POINT_LABELS]
        if 'UNIFORMITY' not in result_df.columns or \
                not all(c in result_df.columns for c in pt_cols):
            return

        best_idx = result_df['UNIFORMITY'].idxmin()
        worst_idx = result_df['UNIFORMITY'].idxmax()

        pts = np.array(WAFER_POINTS_25)
        xi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 250)
        yi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 250)
        XI, YI = np.meshgrid(xi, yi)
        mask = XI**2 + YI**2 > WAFER_RADIUS**2

        fig, axes = plt.subplots(1, 2, figsize=(12, 5.5))
        cases = [(best_idx, 'Best (최소 불균일)', '#4CAF50'),
                 (worst_idx, 'Worst (최대 불균일)', '#F44336')]

        all_vals = result_df[pt_cols].values.flatten()
        vmin = np.nanpercentile(all_vals, 1)
        vmax = np.nanpercentile(all_vals, 99)

        for ax, (ridx, label, color) in zip(axes, cases):
            row = result_df.loc[ridx]
            vals = np.array([row[c] for c in pt_cols], dtype=float)
            run_no = int(row.get('Run', ridx + 1))

            try:
                ZI = griddata(pts, vals, (XI, YI), method='cubic')
            except Exception:
                ZI = griddata(pts, vals, (XI, YI), method='nearest')
            ZI[mask] = np.nan

            im = ax.imshow(ZI, extent=[-WAFER_RADIUS, WAFER_RADIUS,
                                        -WAFER_RADIUS, WAFER_RADIUS],
                           origin='lower', cmap='RdYlGn',
                           vmin=vmin, vmax=vmax, aspect='equal')
            circle = plt.Circle((0, 0), WAFER_RADIUS,
                                 fill=False, color='black', linewidth=2)
            ax.add_patch(circle)

            # 포인트별 값 레이블
            for (px, py), lbl, v in zip(WAFER_POINTS_25, POINT_LABELS, vals):
                ax.annotate(f'{v:.0f}', (px, py),
                            fontsize=6, ha='center', va='center',
                            color='black',
                            bbox=dict(boxstyle='round,pad=0.15', fc='white',
                                      alpha=0.6, ec='none'))

            unif = (vals.max() - vals.min()) / (2 * vals.mean() + 1e-9) * 100
            ax.set_title(f'Run {run_no} — {label}\n'
                         f'Mean={vals.mean():.0f} Å   '
                         f'Uniformity={unif:.2f}%   '
                         f'Max={vals.max():.0f}  Min={vals.min():.0f}',
                         fontsize=9.5, fontweight='bold', color=color)
            ax.set_xlim(-WAFER_RADIUS * 1.05, WAFER_RADIUS * 1.05)
            ax.set_ylim(-WAFER_RADIUS * 1.05, WAFER_RADIUS * 1.05)
            ax.set_xlabel('X (mm)', fontsize=9)
            ax.set_ylabel('Y (mm)', fontsize=9)
            plt.colorbar(im, ax=ax, label='Thickness (Å)')

        fig.suptitle('Best vs Worst Uniformity — 300mm 웨이퍼 두께 분포',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, 'wafer_best_vs_worst.png')

    def plot_wafer_point_stats(self, result_df: pd.DataFrame):
        """25포인트별 두께 통계 (평균 ± σ 및 웨이퍼 위치 히트맵)"""
        from scipy.interpolate import griddata

        pt_cols = [f'T_{lbl}' for lbl in POINT_LABELS]
        if not all(c in result_df.columns for c in pt_cols):
            return

        means = result_df[pt_cols].mean().values
        stds  = result_df[pt_cols].std().values
        pts   = np.array(WAFER_POINTS_25)

        fig, axes = plt.subplots(1, 3, figsize=(16, 5))

        xi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 200)
        yi = np.linspace(-WAFER_RADIUS, WAFER_RADIUS, 200)
        XI, YI = np.meshgrid(xi, yi)
        mask = XI**2 + YI**2 > WAFER_RADIUS**2

        for ax, values, cmap, label in [
            (axes[0], means,         'RdYlGn', 'Mean Thickness (Å)'),
            (axes[1], stds,          'YlOrRd',  '1σ Thickness (Å)'),
            (axes[2], stds/means*100,'YlOrRd',  'CoV (%)'),
        ]:
            try:
                ZI = griddata(pts, values, (XI, YI), method='cubic')
            except Exception:
                ZI = griddata(pts, values, (XI, YI), method='nearest')
            ZI[mask] = np.nan

            im = ax.imshow(ZI, extent=[-WAFER_RADIUS, WAFER_RADIUS,
                                        -WAFER_RADIUS, WAFER_RADIUS],
                           origin='lower', cmap=cmap, aspect='equal')
            circle = plt.Circle((0, 0), WAFER_RADIUS,
                                 fill=False, color='black', linewidth=1.5)
            ax.add_patch(circle)
            ax.scatter(pts[:, 0], pts[:, 1], c='black', s=15, zorder=5)
            for (px, py), v in zip(pts, values):
                ax.text(px, py + 6, f'{v:.1f}', fontsize=5.5,
                        ha='center', va='bottom', color='black')
            ax.set_title(label, fontsize=10, fontweight='bold')
            ax.set_xlabel('X (mm)', fontsize=8)
            ax.set_ylabel('Y (mm)', fontsize=8)
            ax.set_xticks([-100, 0, 100])
            ax.set_yticks([-100, 0, 100])
            plt.colorbar(im, ax=ax, label=label, fraction=0.046)

        fig.suptitle('25포인트 계측 통계 (전체 DOE 런)',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, 'wafer_point_statistics.png')

    # ── 9. 상관관계 히트맵 ───────────────────────────────────────────────────
    def plot_correlation_heatmap(self, data: pd.DataFrame, factors: list, responses: list):
        cols = [c for c in factors + responses if c in data.columns]
        if len(cols) < 2:
            return
        corr = data[cols].corr()
        mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
        fig, ax = plt.subplots(figsize=(max(8, len(cols)), max(7, len(cols) - 1)))
        sns.heatmap(corr, annot=True, fmt='.2f', cmap='RdBu_r', center=0,
                    vmin=-1, vmax=1, ax=ax, square=True, linewidths=0.5,
                    annot_kws={'size': 8})
        ax.set_title('Correlation Heatmap (Factors & Responses)',
                     fontsize=13, fontweight='bold')
        fig.tight_layout()
        self._save(fig, 'correlation_heatmap.png')


# ══════════════════════════════════════════════════════════════════════════════
#  보고서 생성기
# ══════════════════════════════════════════════════════════════════════════════
class DOEReporter:

    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def save_excel(self, design_df: pd.DataFrame, result_df: pd.DataFrame,
                   analysis_results: dict, opt_result: dict):
        path = os.path.join(self.output_dir, 'doe_report.xlsx')
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            design_df.to_excel(writer, sheet_name='실험계획', index=False)
            # 응답 및 비포인트 컬럼만
            main_cols = [c for c in result_df.columns if not c.startswith('T_P')]
            result_df[main_cols].to_excel(writer, sheet_name='실험결과', index=False)
            # 25포인트 두께 데이터
            pt_cols = ['Run'] + [f'T_{lbl}' for lbl in POINT_LABELS
                                 if f'T_{lbl}' in result_df.columns]
            if len(pt_cols) > 1:
                pt_df = result_df[pt_cols].copy()
                # 좌표 정보 헤더 행 추가
                coord_row = {'Run': 'X(mm)'}
                coord_row.update({f'T_{lbl}': WAFER_POINTS_25[i][0]
                                  for i, lbl in enumerate(POINT_LABELS)
                                  if f'T_{lbl}' in result_df.columns})
                coord_row2 = {'Run': 'Y(mm)'}
                coord_row2.update({f'T_{lbl}': WAFER_POINTS_25[i][1]
                                   for i, lbl in enumerate(POINT_LABELS)
                                   if f'T_{lbl}' in result_df.columns})
                header_df = pd.DataFrame([coord_row, coord_row2])
                pd.concat([header_df, pt_df], ignore_index=True).to_excel(
                    writer, sheet_name='25pt_두께', index=False
                )
            for resp, adict in analysis_results.items():
                if 'anova' in adict and not adict['anova'].empty:
                    adict['anova'].to_excel(writer, sheet_name=f'ANOVA_{resp}')
                if 'main_effects' in adict and not adict['main_effects'].empty:
                    adict['main_effects'].to_excel(writer,
                                                    sheet_name=f'MainEff_{resp}', index=False)
                if 'interactions' in adict and not adict['interactions'].empty:
                    adict['interactions'].to_excel(writer,
                                                   sheet_name=f'Interact_{resp}', index=False)
            # 최적화 결과
            opt_df = pd.DataFrame([{
                'Item': 'Desirability', 'Value': opt_result['desirability'],
            }] + [{'Item': f'FACTOR_{k}', 'Value': v}
                  for k, v in opt_result['optimal_factors'].items()] +
                  [{'Item': f'PRED_{k}', 'Value': v}
                   for k, v in opt_result['predicted_responses'].items()])
            opt_df.to_excel(writer, sheet_name='최적화결과', index=False)
        print(f"  엑셀 저장: {path}")
        return path

    def print_summary(self, design_df: pd.DataFrame, analysis_results: dict,
                      opt_result: dict):
        sep = '=' * 65
        print(f'\n{sep}')
        print('  반도체 공정 DOE 분석 결과 요약')
        print(f'  분석 일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        print(sep)

        print(f'\n[실험 설계]')
        print(f'  총 실험 횟수: {len(design_df)} 회')
        print(f'  인자 수    : {len([c for c in design_df.columns if c in FACTORS])} 개')

        for resp, adict in analysis_results.items():
            print(f'\n[{resp} 분석]')
            me = adict.get('main_effects', pd.DataFrame())
            if not me.empty:
                sig = me[me['Significant']]
                print(f'  유의한 주효과 ({len(sig)}개):',
                      ', '.join(sig['Factor'].tolist()) if not sig.empty else '없음')
            ie = adict.get('interactions', pd.DataFrame())
            if not ie.empty:
                top3 = ie.head(3)
                print(f'  주요 교호작용 Top3:')
                for _, r in top3.iterrows():
                    print(f'    {r["Factors"]}: {r["Interaction_Effect"]:+.3f}')
            reg = adict.get('regression', {})
            if reg:
                print(f'  회귀 R²: {reg["r2"]:.4f}')

        print(f'\n[최적화 결과]')
        print(f'  종합 만족도 (Desirability): {opt_result["desirability"]:.4f}')
        print(f'\n  최적 인자 조건:')
        for f, v in opt_result['optimal_factors'].items():
            unit = FACTORS.get(f, {}).get('unit', '')
            print(f'    {f:10s}: {v:8.2f} {unit}')
        print(f'\n  예측 결과:')
        for r, v in opt_result['predicted_responses'].items():
            unit = RESPONSES.get(r, {}).get('unit', '')
            target = RESPONSES.get(r, {}).get('target', '?')
            print(f'    {r:12s}: {v:8.2f} {unit}  (Target: {target})')
        print(f'\n{sep}\n')


# ══════════════════════════════════════════════════════════════════════════════
#  데이터 시뮬레이터 (실제 데이터가 없을 때 사용)
# ══════════════════════════════════════════════════════════════════════════════
class ProcessSimulator:
    """반도체 공정 응답 시뮬레이터 (물리 기반 근사 모델)"""

    def __init__(self, noise_level: float = 0.05):
        self.noise = noise_level
        np.random.seed(42)

    def _noise(self, value: float) -> float:
        return value * (1 + np.random.normal(0, self.noise))

    def thickness(self, params: dict) -> float:
        """두께 모델 (C3H6, TIME, PRESSURE, RF_POWER 주요 영향)"""
        c3h6 = params.get('C3H6', 0)
        time_ = params.get('TIME', 60)
        pressure = params.get('PRESSURE', 5)
        n2 = params.get('N2', 100)
        ar = params.get('AR', 50)
        space = params.get('SPACE', 15)
        rf = params.get('RF_POWER', 1000)
        # 기본 증착률 (RF 파워가 높을수록 플라즈마 밀도↑ → 증착률 증가)
        dep_rate = 40 + 0.8 * c3h6 + 0.5 * n2 * 0.1 - 0.3 * ar * 0.1
        dep_rate *= (pressure / 5) ** 0.6
        dep_rate *= max(0.5, 1 - (space - 15) * 0.015)
        dep_rate *= (rf / 1000) ** 0.45
        thick = dep_rate * time_
        # 교호작용
        thick += 2.5 * (c3h6 - 25) * (pressure - 5) * 0.1
        thick -= 1.0 * (time_ - 60) * (space - 15) * 0.02
        thick += 0.8 * (rf - 1000) * (c3h6 - 25) * 0.0001
        return max(0, self._noise(thick))

    def uniformity(self, params: dict) -> float:
        """균일도 모델 (SPACE, HE, AR, RF_POWER 주요 영향, 낮을수록 좋음)"""
        space = params.get('SPACE', 15)
        he = params.get('HE', 250)
        ar = params.get('AR', 50)
        pressure = params.get('PRESSURE', 5)
        nf3 = params.get('NF3', 50)
        rf = params.get('RF_POWER', 1000)
        uni = 5.0
        uni -= 0.08 * (he - 250) * 0.01
        uni += 0.05 * abs(space - 15) * 0.2
        uni += 0.03 * pressure
        uni -= 0.02 * ar * 0.01
        uni += 0.01 * nf3 * 0.01
        # RF 파워: 적정 파워에서 균일도 최적 (너무 낮거나 높으면 불균일)
        uni += 0.0015 * abs(rf - 1000)
        # 교호작용
        uni += 0.005 * (he - 250) * 0.01 * (space - 15)
        uni -= 0.0001 * (rf - 1000) * (he - 250) * 0.001
        uni = max(0.1, abs(self._noise(uni)))
        return min(uni, 15.0)

    def particle(self, params: dict) -> float:
        """파티클 모델 (NF3, O2 클리닝 효과, RF_POWER, 낮을수록 좋음)"""
        nf3 = params.get('NF3', 50)
        o2 = params.get('O2', 50)
        time_ = params.get('TIME', 60)
        pressure = params.get('PRESSURE', 5)
        c3h6 = params.get('C3H6', 25)
        rf = params.get('RF_POWER', 1000)
        base_particle = 30
        base_particle -= 0.2 * nf3
        base_particle -= 0.15 * o2
        base_particle += 0.05 * c3h6
        base_particle += 0.1 * pressure
        base_particle += 0.02 * time_
        # RF 파워가 높을수록 파티클 증가 (아크 방전, 챔버 벽 스퍼터링)
        base_particle += 0.008 * (rf - 1000)
        # 교호작용 (NF3*O2 시너지, RF_POWER*PRESSURE)
        base_particle -= 0.002 * nf3 * o2
        base_particle += 0.0005 * (rf - 1000) * (pressure - 5)
        particle = max(0, self._noise(base_particle))
        return round(particle)

    def _spatial_factor(self, x: float, y: float, params: dict) -> float:
        """
        웨이퍼 위치 (x, y) mm 에서의 두께 공간 변화 계수 반환.
        공정 조건에 따라 bowl / dome / edge 프로파일을 모사.
        """
        R = WAFER_RADIUS
        rn = np.sqrt(x**2 + y**2) / R   # 정규화 반경 [0, 1]
        theta = np.arctan2(y, x)

        space = params.get('SPACE', 15)
        he    = params.get('HE', 250)
        rf    = params.get('RF_POWER', 1000)
        pressure = params.get('PRESSURE', 5)
        n2    = params.get('N2', 100)

        # SPACE: 간격 작으면 중심 두꺼운 dome, 크면 가장자리 두꺼운 bowl
        space_coeff = (space - 15) / 25
        radial = space_coeff * 0.12 * rn**2

        # HE: 플라즈마 균일화 효과 → 반경 기울기 감소
        he_smooth = max(0.2, 1 - (he / 500) * 0.45)
        radial *= he_smooth

        # RF_POWER: 고출력 → 가장자리 가열 → 가장자리 박막화
        rf_coeff = (rf - 1000) / 1900
        radial += rf_coeff * 0.07 * (rn**2 - 0.25)

        # PRESSURE: 높을수록 가장자리 소모 증가
        radial += (pressure - 5.5) / 9.0 * 0.05 * rn

        # N2 흐름 방향 비대칭 (작은 각도 variation)
        n2_asym = (n2 - 100) / 200 * 0.015
        angular = n2_asym * np.sin(theta) * rn

        return 1.0 + radial + angular

    def simulate(self, design_df: pd.DataFrame) -> pd.DataFrame:
        """전체 실험 계획에 대해 응답 시뮬레이션 (25포인트 계측)"""
        df = design_df.copy()
        pt_cols = {lbl: [] for lbl in POINT_LABELS}
        thickness_vals, unif_vals, particle_vals = [], [], []

        for _, row in df.iterrows():
            params = {}
            for f in FACTORS:
                actual_col = f'{f}_actual'
                if actual_col in row.index:
                    params[f] = row[actual_col]
                elif f in row.index:
                    lo, hi = FACTORS[f]['low'], FACTORS[f]['high']
                    params[f] = lo + (row[f] + 1) / 2 * (hi - lo)

            # 평균 두께 계산
            mean_thick = self.thickness(params)

            # 25포인트별 두께: 공간 변화 계수 × 평균 두께 + 측정 노이즈
            pt_thicknesses = []
            for (px, py), lbl in zip(WAFER_POINTS_25, POINT_LABELS):
                sf = self._spatial_factor(px, py, params)
                t = mean_thick * sf * (1 + np.random.normal(0, self.noise * 0.3))
                t = max(0, t)
                pt_cols[lbl].append(t)
                pt_thicknesses.append(t)

            pt_arr = np.array(pt_thicknesses)
            pt_mean = pt_arr.mean()
            # UNIFORMITY = (max-min) / (2*mean) × 100 (%)
            unif = (pt_arr.max() - pt_arr.min()) / (2 * pt_mean + 1e-9) * 100

            thickness_vals.append(pt_mean)
            unif_vals.append(unif)
            particle_vals.append(self.particle(params))

        for lbl in POINT_LABELS:
            df[f'T_{lbl}'] = pt_cols[lbl]
        df['THICKNESS']  = thickness_vals
        df['UNIFORMITY'] = unif_vals
        df['PARTICLE']   = particle_vals
        return df


# ══════════════════════════════════════════════════════════════════════════════
#  메인 파이프라인
# ══════════════════════════════════════════════════════════════════════════════
def run_doe_analysis(
    active_factors: list = None,
    design_type: str = 'plackett_burman',
    data_csv: str = None,
    output_dir: str = OUTPUT_DIR,
):
    """
    DOE 분석 전체 파이프라인

    Parameters
    ----------
    active_factors : 분석에 사용할 인자 목록 (None=전체)
    design_type    : 'full_factorial' | 'ccd' | 'box_behnken' | 'plackett_burman'
    data_csv       : 실제 측정 데이터 CSV 경로 (None=시뮬레이션)
    output_dir     : 결과 저장 폴더
    """
    os.makedirs(output_dir, exist_ok=True)
    print('\n' + '=' * 65)
    print('  반도체 공정 DOE 분석 시스템  v1.0')
    print('  목표: 두께 / 균일도 / 파티클 최적화')
    print('=' * 65)

    # ── 1. 실험 설계 ────────────────────────────────────────────────────────
    print(f'\n[1단계] 실험 설계 생성 ({design_type})')
    used_factors = active_factors or list(FACTORS.keys())
    designer = DOEDesigner(FACTORS, used_factors)
    design_map = {
        'full_factorial': designer.full_factorial_2level,
        'ccd': designer.central_composite,
        'box_behnken': designer.box_behnken,
        'plackett_burman': designer.plackett_burman,
    }
    design_fn = design_map.get(design_type, designer.plackett_burman)
    design_df = design_fn()
    print(f'  실험 횟수: {len(design_df)} runs')
    print(f'  사용 인자: {", ".join(used_factors)}')

    # ── 2. 결과 데이터 로드 또는 시뮬레이션 ─────────────────────────────────
    print('\n[2단계] 실험 결과 데이터 준비')
    if data_csv and os.path.exists(data_csv):
        result_df = pd.read_csv(data_csv)
        print(f'  실제 데이터 로드: {data_csv}')
    else:
        print('  시뮬레이션 데이터 생성 중...')
        sim = ProcessSimulator(noise_level=0.04)
        result_df = sim.simulate(design_df)
        print(f'  시뮬레이션 완료: {len(result_df)}개 실험')

    # 통계 요약 출력
    for resp in RESPONSES:
        if resp in result_df.columns:
            col = result_df[resp]
            print(f'    {resp:12s}: mean={col.mean():.2f}, '
                  f'std={col.std():.2f}, '
                  f'min={col.min():.2f}, max={col.max():.2f}')

    # ── 3. 통계 분석 ─────────────────────────────────────────────────────────
    print('\n[3단계] 통계 분석 (주효과, 교호작용, 회귀)')
    resp_list = [r for r in RESPONSES if r in result_df.columns]
    analyzer = DOEAnalyzer(result_df, used_factors, resp_list)
    analysis_results = {}
    for resp in resp_list:
        print(f'  >> {resp} 분석 중...')
        analysis_results[resp] = {
            'anova': analyzer.anova(resp, interactions=2),
            'main_effects': analyzer.main_effects(resp),
            'interactions': analyzer.interaction_effects(resp),
            'pareto': analyzer.pareto_effects(resp),
            'regression': analyzer.regression(resp, degree=2),
        }

    # ── 4. 최적화 ────────────────────────────────────────────────────────────
    print('\n[4단계] 다중 반응 최적화 (Desirability Function)')
    optimizer = DOEOptimizer(analyzer, RESPONSES)
    opt_result = optimizer.grid_search(n_grid=8)

    # ── 5. 시각화 ────────────────────────────────────────────────────────────
    print(f'\n[5단계] 시각화 생성 → {output_dir}/')
    viz = DOEVisualizer(output_dir)
    # 설계 행렬
    viz.plot_design_matrix(design_df, used_factors,
                           f'Design Matrix ({design_type})')
    # 응답별 그래프
    for resp in resp_list:
        viz.plot_main_effects(result_df, used_factors, resp)
        viz.plot_interactions(result_df, used_factors, resp, top_n=6)
        viz.plot_pareto(analysis_results[resp]['pareto'], resp)
        viz.plot_residuals(analyzer, resp)
        # 반응 표면 (주요 인자 쌍)
        me = analysis_results[resp]['main_effects']
        if not me.empty and len(me) >= 2:
            top2 = me.head(2)['Factor'].tolist()
            viz.plot_response_surface(analyzer, resp, top2[0], top2[1])
    # 상관관계
    viz.plot_correlation_heatmap(result_df, used_factors, resp_list)
    # 최적화 결과
    viz.plot_optimization_result(opt_result, RESPONSES)
    # 웨이퍼 맵 (25포인트)
    viz.plot_wafer_maps(result_df)
    viz.plot_wafer_best_worst(result_df)
    viz.plot_wafer_point_stats(result_df)

    # ── 6. 보고서 저장 ───────────────────────────────────────────────────────
    print('\n[6단계] 보고서 저장')
    reporter = DOEReporter(output_dir)
    reporter.save_excel(design_df, result_df, analysis_results, opt_result)
    reporter.print_summary(design_df, analysis_results, opt_result)

    # 결과 JSON 저장
    result_json = {
        'timestamp': datetime.now().isoformat(),
        'design_type': design_type,
        'active_factors': used_factors,
        'n_runs': len(design_df),
        'optimization': {
            'desirability': opt_result['desirability'],
            'optimal_factors': opt_result['optimal_factors'],
            'predicted_responses': opt_result['predicted_responses'],
        },
    }
    json_path = os.path.join(output_dir, 'doe_result.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result_json, f, ensure_ascii=False, indent=2)
    print(f'  JSON 저장: {json_path}')
    print(f'\n완료. 생성된 파일: {len(viz.saved_files) + 2}개')
    return result_df, analysis_results, opt_result


# ══════════════════════════════════════════════════════════════════════════════
#  Excel 템플릿 생성 & 로드
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_template(
    design_type: str = 'plackett_burman',
    active_factors: list = None,
    output_path: str = 'doe_input_template.xlsx',
) -> str:
    """
    실험 계획이 채워진 Excel 입력 템플릿 생성.
    노란색 셀에 측정값을 입력한 후 --data 옵션으로 로드.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter

    used = active_factors or list(FACTORS.keys())
    designer = DOEDesigner(FACTORS, used)
    design_map = {
        'full_factorial':   designer.full_factorial_2level,
        'ccd':              designer.central_composite,
        'box_behnken':      designer.box_behnken,
        'plackett_burman':  designer.plackett_burman,
    }
    design_df = design_map.get(design_type, designer.plackett_burman)()
    n_runs = len(design_df)

    wb = Workbook()

    # ── 색상 정의 ────────────────────────────────────────────────────────────
    BLUE_FILL   = PatternFill('solid', fgColor='1565C0')
    GRAY_FILL   = PatternFill('solid', fgColor='CFD8DC')
    YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
    GREEN_FILL  = PatternFill('solid', fgColor='E8F5E9')
    ORANGE_FILL = PatternFill('solid', fgColor='FFF3E0')
    WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    def hdr(cell, text, fill=BLUE_FILL, font_color='FFFFFF', size=10):
        cell.value = text
        cell.fill  = fill
        cell.font  = Font(bold=True, color=font_color, size=size)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = thin
    def data_cell(cell, value=None, fill=WHITE_FILL, num_fmt=None):
        cell.value  = value
        cell.fill   = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if num_fmt:
            cell.number_format = num_fmt

    # ══════════════════════════════════════════════════════════════════════════
    #  Sheet 1: 측정 데이터 입력
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = '측정데이터입력'
    ws.freeze_panes = 'A3'

    # 행 1: 섹션 타이틀
    actual_cols  = [f'{f}_actual' for f in used
                    if f'{f}_actual' in design_df.columns]
    factor_count = len(actual_cols)
    resp_start   = 2 + factor_count          # Run + factors
    pt_start     = resp_start + 3            # 응답 3개 뒤
    total_cols   = pt_start + 25

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    ws.merge_cells(start_row=1, start_column=2,
                   end_row=1, end_column=1+factor_count)
    ws.merge_cells(start_row=1, start_column=resp_start,
                   end_row=1, end_column=resp_start+2)
    ws.merge_cells(start_row=1, start_column=pt_start,
                   end_row=1, end_column=pt_start+24)

    hdr(ws.cell(1,1), 'Run', GRAY_FILL, '263238')
    hdr(ws.cell(1,2), '공정 조건 (자동 입력 — 수정 금지)', GRAY_FILL, '263238')
    hdr(ws.cell(1,resp_start), '▶ 응답 측정값 (노란색 셀에 입력)', BLUE_FILL)
    hdr(ws.cell(1,pt_start),
        '▶ 두께 25포인트 측정값 (Å) — 입력 시 THICKNESS/UNIFORMITY 자동 계산',
        PatternFill('solid', fgColor='E65100'), 'FFFFFF')

    # 행 2: 컬럼 헤더
    hdr(ws.cell(2,1), 'Run No.', GRAY_FILL, '263238')
    for ci, col in enumerate(actual_cols, start=2):
        f = col.replace('_actual','')
        unit = FACTORS[f]['unit']
        hdr(ws.cell(2, ci), f'{f}\n({unit})', GRAY_FILL, '263238')

    resp_headers = [
        ('THICKNESS (Å)',   '※ 25pt 미입력 시'),
        ('UNIFORMITY (%)',  '※ 25pt 미입력 시'),
        ('PARTICLE (ea)',   '필수 입력'),
    ]
    for ci, (label, note) in enumerate(resp_headers):
        hdr(ws.cell(2, resp_start+ci), f'{label}\n{note}',
            PatternFill('solid', fgColor='F57F17'), '263238')

    for ci, lbl in enumerate(POINT_LABELS):
        x, y = WAFER_POINTS_25[ci]
        r = np.sqrt(x**2+y**2)
        hdr(ws.cell(2, pt_start+ci),
            f'{lbl}\n({x:.0f},{y:.0f})\nr={r:.0f}',
            ORANGE_FILL, '263238', size=8)

    # 행 3~: 데이터
    for ri, (_, row) in enumerate(design_df.iterrows(), start=3):
        run_no = int(row.get('Run', ri-2))
        data_cell(ws.cell(ri, 1), run_no, GRAY_FILL)
        for ci, col in enumerate(actual_cols, start=2):
            v = row[col]
            data_cell(ws.cell(ri, ci), round(v, 3), WHITE_FILL, '0.000')
        for ci in range(3):
            data_cell(ws.cell(ri, resp_start+ci), None, YELLOW_FILL, '0.00')
        for ci in range(25):
            data_cell(ws.cell(ri, pt_start+ci), None, YELLOW_FILL, '0')

    # 열 너비
    ws.column_dimensions['A'].width = 8
    for ci in range(factor_count):
        ws.column_dimensions[get_column_letter(2+ci)].width = 11
    for ci in range(3):
        ws.column_dimensions[get_column_letter(resp_start+ci)].width = 14
    for ci in range(25):
        ws.column_dimensions[get_column_letter(pt_start+ci)].width = 9
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 42
    for ri in range(3, 3+n_runs):
        ws.row_dimensions[ri].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    #  Sheet 2: 25포인트 좌표 안내
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('25포인트_좌표')
    headers2 = ['Point', 'X (mm)', 'Y (mm)', 'Radius (mm)', 'Angle (°)', '위치 설명']
    for ci, h in enumerate(headers2, 1):
        hdr(ws2.cell(1,ci), h)
    zone_map = {0:'Center', 1:'Inner(35mm)', 2:'Inner(35mm)', 3:'Inner(35mm)', 4:'Inner(35mm)',
                5:'Mid-diag(75mm)', 6:'Mid-diag(75mm)', 7:'Mid-diag(75mm)', 8:'Mid-diag(75mm)',
                9:'Mid-card(105mm)',10:'Mid-card(105mm)',11:'Mid-card(105mm)',12:'Mid-card(105mm)',
               13:'Out-diag(120mm)',14:'Out-diag(120mm)',15:'Out-diag(120mm)',16:'Out-diag(120mm)'}
    for i, (lbl,(x,y)) in enumerate(zip(POINT_LABELS, WAFER_POINTS_25)):
        r    = np.sqrt(x**2+y**2)
        ang  = np.degrees(np.arctan2(y,x))
        zone = zone_map.get(i, f'Edge(140mm)')
        row  = [lbl, x, y, round(r,1), round(ang,1), zone]
        for ci, v in enumerate(row, 1):
            data_cell(ws2.cell(i+2,ci), v,
                      GREEN_FILL if i==0 else WHITE_FILL)
    for ci, w in enumerate([8,10,10,12,12,18],1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    #  Sheet 3: 작성 방법 안내
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('작성방법안내')
    guide = [
        ('반도체 공정 DOE 데이터 입력 가이드', True, 'FFFFFF', '1565C0', 16),
        ('', False, '000000', 'FFFFFF', 11),
        ('■ 기본 사용법', True, '000000', 'E3F2FD', 12),
        ('1. [측정데이터입력] 시트의 노란색(▶) 셀에만 값을 입력하세요.', False,'000000','FFFFFF',11),
        ('2. 회색 셀(공정 조건)은 자동 계산값이므로 수정하지 마세요.', False,'000000','FFFFFF',11),
        ('3. 저장 후 아래 명령어로 분석을 실행하세요.', False,'000000','FFFFFF',11),
        ('', False, '000000', 'FFFFFF', 11),
        ('■ 실행 명령어', True, '000000', 'E3F2FD', 12),
        ('  python3 doe_html_report.py --data doe_input_template.xlsx', False,'E65100','FFF9C4',11),
        ('', False, '000000', 'FFFFFF', 11),
        ('■ 입력 우선순위', True, '000000', 'E3F2FD', 12),
        ('방법 A (권장): 25포인트 두께(T_P01~T_P25) 모두 입력', False,'000000','FFFFFF',11),
        ('  → THICKNESS 와 UNIFORMITY 는 자동 계산됩니다', False,'000000','FFFFFF',11),
        ('  → PARTICLE(ea) 은 별도 입력 필요', False,'000000','FFFFFF',11),
        ('방법 B: THICKNESS, UNIFORMITY, PARTICLE 직접 입력', False,'000000','FFFFFF',11),
        ('  → 25포인트 열은 비워도 됩니다', False,'000000','FFFFFF',11),
        ('', False, '000000', 'FFFFFF', 11),
        ('■ 25포인트 계측 순서 (300mm 웨이퍼, EE=3mm)', True,'000000','E3F2FD',12),
        ('  P01: 중심 (0, 0)', False,'000000','FFFFFF',11),
        ('  P02~P05: r=35mm 상하좌우', False,'000000','FFFFFF',11),
        ('  P06~P09: r=75mm 대각', False,'000000','FFFFFF',11),
        ('  P10~P13: r=105mm 상하좌우', False,'000000','FFFFFF',11),
        ('  P14~P17: r=120mm 대각', False,'000000','FFFFFF',11),
        ('  P18~P25: r=140mm 8방향', False,'000000','FFFFFF',11),
        ('', False,'000000','FFFFFF',11),
        ('■ UNIFORMITY 계산식', True,'000000','E3F2FD',12),
        ('  (MAX - MIN) / (2 × MEAN) × 100  (%)', False,'1565C0','FFFFFF',11),
    ]
    for ri, (text, bold, fcolor, bgcolor, fsize) in enumerate(guide, 1):
        c = ws3.cell(ri, 1, text)
        c.font   = Font(bold=bold, color=fcolor, size=fsize)
        c.fill   = PatternFill('solid', fgColor=bgcolor)
        c.alignment = Alignment(vertical='center')
        ws3.row_dimensions[ri].height = 20
    ws3.column_dimensions['A'].width = 70

    wb.save(output_path)
    print(f'  Excel 템플릿 생성: {output_path}')
    print(f'  → [{n_runs}개 Run] 노란색 셀에 측정값 입력 후 저장하세요.')
    return output_path


def load_excel_data(path: str) -> pd.DataFrame:
    """
    generate_excel_template() 으로 생성한 Excel 파일에서
    측정 데이터를 읽어 result_df 형태로 반환.
    25포인트가 입력되면 THICKNESS/UNIFORMITY 자동 계산.
    """
    df = pd.read_excel(path, sheet_name='측정데이터입력', header=1)

    # 헤더 정규화: 멀티라인, 단위 제거 → 핵심 키워드만 남김
    rename_map = {}
    for col in df.columns:
        raw = str(col).split('\n')[0].strip()
        # "Run No." → "Run"
        if 'Run' in raw:
            rename_map[col] = 'Run'
            continue
        # "THICKNESS (Å)\n..." → "THICKNESS"
        for resp in ['THICKNESS', 'UNIFORMITY', 'PARTICLE']:
            if resp in raw.upper():
                rename_map[col] = resp
                break
        else:
            # 인자: "AR\n(sccm)" → "AR_raw"
            for f in FACTORS:
                if raw.startswith(f) or raw == f:
                    rename_map[col] = f'{f}_raw'
                    break
            else:
                # 25포인트: "P01\n(0,0)\nr=0" → "T_P01"
                for lbl in POINT_LABELS:
                    if lbl in raw:
                        rename_map[col] = f'T_{lbl}'
                        break
    df = df.rename(columns=rename_map)

    # 25포인트가 채워진 경우 THICKNESS/UNIFORMITY 재계산
    pt_cols_present = [f'T_{lbl}' for lbl in POINT_LABELS
                       if f'T_{lbl}' in df.columns]
    if len(pt_cols_present) == 25:
        pt_data = df[pt_cols_present].apply(
            pd.to_numeric, errors='coerce').values.astype(float)
        filled = ~np.isnan(pt_data).all(axis=1)
        if filled.any():
            safe = np.where(np.isnan(pt_data), 0, pt_data)
            means = safe.mean(axis=1)
            maxs  = safe.max(axis=1)
            mins  = safe.min(axis=1)
            # 25포인트가 채워진 행만 덮어쓰기
            df.loc[filled, 'THICKNESS']  = means[filled]
            df.loc[filled, 'UNIFORMITY'] = (
                (maxs[filled] - mins[filled]) /
                (2 * means[filled] + 1e-9) * 100
            )

    # 인자 coded/actual 컬럼 복원
    for f in FACTORS:
        raw_col = f'{f}_raw'
        if raw_col in df.columns:
            df[f'{f}_actual'] = pd.to_numeric(df[raw_col], errors='coerce')
            lo, hi = FACTORS[f]['low'], FACTORS[f]['high']
            df[f] = 2 * (df[f'{f}_actual'] - lo) / (hi - lo) - 1

    # 응답 숫자 변환
    for col in ['THICKNESS', 'UNIFORMITY', 'PARTICLE']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    df = df.dropna(subset=['Run'])
    df['Run'] = df['Run'].astype(int)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  데이터 입력 인터페이스
# ══════════════════════════════════════════════════════════════════════════════
def input_actual_data(design_df: pd.DataFrame) -> pd.DataFrame:
    """실제 실험 결과를 수동으로 입력"""
    print('\n실제 측정 데이터 입력')
    print('각 Run의 두께(Å), 균일도(%), 파티클(ea) 을 입력하세요.')
    print('(입력 없이 Enter = 시뮬레이션 값 사용)\n')
    df = design_df.copy()
    sim = ProcessSimulator(noise_level=0.03)
    sim_df = sim.simulate(design_df)
    thickness_vals, unif_vals, particle_vals = [], [], []
    for i, row in df.iterrows():
        run_no = row.get('Run', i+1)
        print(f'  Run {run_no:2d}:', end=' ')
        for f in list(FACTORS.keys())[:4]:
            actual_col = f'{f}_actual'
            if actual_col in row.index:
                print(f'{f}={row[actual_col]:.0f}', end=' ')
        print()
        try:
            t_in = input(f'    두께 (Å) [default={sim_df.loc[i,"THICKNESS"]:.0f}]: ').strip()
            u_in = input(f'    균일도 (%) [default={sim_df.loc[i,"UNIFORMITY"]:.2f}]: ').strip()
            p_in = input(f'    파티클 (ea) [default={sim_df.loc[i,"PARTICLE"]:.0f}]: ').strip()
            thickness_vals.append(float(t_in) if t_in else sim_df.loc[i, 'THICKNESS'])
            unif_vals.append(float(u_in) if u_in else sim_df.loc[i, 'UNIFORMITY'])
            particle_vals.append(int(p_in) if p_in else sim_df.loc[i, 'PARTICLE'])
        except (ValueError, KeyboardInterrupt):
            thickness_vals.append(sim_df.loc[i, 'THICKNESS'])
            unif_vals.append(sim_df.loc[i, 'UNIFORMITY'])
            particle_vals.append(sim_df.loc[i, 'PARTICLE'])
    df['THICKNESS'] = thickness_vals
    df['UNIFORMITY'] = unif_vals
    df['PARTICLE'] = particle_vals
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  엔트리포인트
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='반도체 공정 DOE 분석 시스템'
    )
    parser.add_argument(
        '--design', choices=['full_factorial', 'ccd', 'box_behnken', 'plackett_burman'],
        default='plackett_burman',
        help='실험 설계 유형 (기본값: plackett_burman)',
    )
    parser.add_argument(
        '--factors', nargs='+', choices=list(FACTORS.keys()),
        default=None,
        help='분석할 인자 목록 (기본값: 전체)',
    )
    parser.add_argument(
        '--data', default=None,
        help='실제 측정 데이터 파일 경로 (.csv 또는 .xlsx)',
    )
    parser.add_argument(
        '--output', default=OUTPUT_DIR,
        help='결과 저장 폴더',
    )
    parser.add_argument(
        '--template', action='store_true',
        help='Excel 입력 템플릿 생성 후 종료 (측정값 입력용)',
    )
    parser.add_argument(
        '--interactive', action='store_true',
        help='실제 실험 데이터 수동 입력 모드',
    )
    args = parser.parse_args()

    if args.template:
        used = args.factors or list(FACTORS.keys())
        os.makedirs(args.output, exist_ok=True)
        tpl_path = os.path.join(args.output, 'doe_input_template.xlsx')
        generate_excel_template(
            design_type=args.design,
            active_factors=used,
            output_path=tpl_path,
        )
        print(f'\n사용법:')
        print(f'  1. {tpl_path} 를 엑셀로 열기')
        print(f'  2. 노란색 셀에 측정값 입력 후 저장')
        print(f'  3. python3 doe_semiconductor.py --data {tpl_path}')
        import sys; sys.exit(0)

    # xlsx 파일 지원
    data_path = args.data
    if data_path and data_path.endswith('.xlsx'):
        print(f'  Excel 데이터 로드: {data_path}')
        loaded_df = load_excel_data(data_path)
        tmp_csv = data_path.replace('.xlsx', '_loaded.csv')
        loaded_df.to_csv(tmp_csv, index=False)
        data_path = tmp_csv

    if args.interactive:
        # 인터랙티브 입력 모드
        used = args.factors or list(FACTORS.keys())
        designer = DOEDesigner(FACTORS, used)
        design_df = designer.plackett_burman()
        result_df = input_actual_data(design_df)
        tmp_csv = os.path.join(args.output, '_input_data.csv')
        os.makedirs(args.output, exist_ok=True)
        result_df.to_csv(tmp_csv, index=False)
        run_doe_analysis(
            active_factors=used,
            design_type=args.design,
            data_csv=tmp_csv,
            output_dir=args.output,
        )
    else:
        run_doe_analysis(
            active_factors=args.factors,
            design_type=args.design,
            data_csv=data_path,
            output_dir=args.output,
        )
